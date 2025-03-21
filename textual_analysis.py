# BEFORE EXECUTING THIS FILE YOU HAVE TO ENSURE THAT YOU HAVE SOME NLTK PACKAGES DOWNLOADED IN YOUR SYS IF NOT THEN REVIEW THE COMMENTS I HAVE LEFT.
# AND STORE ALL EXTRACTED TEXT FILES IN FOLDER 'ext_txt'

import os
import nltk
import re
import openpyxl
import pandas as pd

# nltk.download('vader_lexicon')
# nltk.download('punkt')
# nltk.download('cmudict')

# SENTIMENT ANALYSIS

# Load positive and negative word lists
pos_words = open('positive-words.txt', 'r').read().split('\n')
neg_words = open('negative-words.txt', 'r').read().split('\n')

# Load stopwords
stop_words = open('StopWords_GenericLong.txt', 'r').read().split('\n')

# Load CMU dictionary
cmudict = nltk.corpus.cmudict.dict()

def analyze_sentiment(text,url_id):

    # Initialize sentiment scores and readability variables
    positive_score = 0
    negative_score = 0
    neutral_score = 0
    total_words = 0
    total_sentences = 0
    complex_words = 0
    before_words = 0
    syllables = 0
    personal_pronouns = 0
    total_chars =0

    # Tokenize text into individual words and sentences
    words = re.findall('\w+', text.lower())
    sentences = nltk.sent_tokenize(text)

    # Loop through each word
    for word in words:
        if word in ['i', 'we', 'my', 'ours', 'us']:
            personal_pronouns += 1

        # Ignore stopwords
        if word in stop_words or word in ['?', '!', ',', '.']:
            continue
        
        # Count the number of syllables in the word
        vowels = 'aeiou'
        num_syllables = 0
        for i in range(len(word)):
            if i == 0 and word[i] in vowels:
                num_syllables += 1
            elif i > 0 and word[i] in vowels and word[i-1] not in vowels:
                num_syllables += 1
        if word.endswith(('es', 'ed')) and num_syllables > 1:
            num_syllables -= 1
        syllables += num_syllables
        
        # Check if word is positive, negative, or neutral
        if word in pos_words:
            positive_score += 1
        elif word in neg_words:
            negative_score += 1
        else:
            neutral_score += 1

        # Check if word is a personal pronoun
        total_words += 1
        total_chars += len(word)
        avg_word_length = total_chars / total_words

        # Check if word is complex
        if len(word) >= 7 and word.lower() in cmudict:
            complex_words += 1


    # Calculate polarity and subjectivity scores
    polarity_score = (positive_score - negative_score) / (positive_score + negative_score + 0.000001)
    subjectivity_score = (positive_score + negative_score) / (total_words + 0.000001)

    # Calculate readability score
    total_sentences = len(sentences)
    average_sentence_length = total_words / total_sentences
    percentage_complex_words = (complex_words / total_words) * 100
    fog_index = 0.4 * (average_sentence_length + percentage_complex_words)
    avg_num_words_per_sentence = total_words / total_sentences
    avg_syllables_per_word = syllables / total_words
    
    # Read the Excel file into a pandas dataframe
    df = pd.read_excel("output.xlsx")
    row = df.loc[df['URL_ID'] == url_id].index[0] +2

    # Load the workbook
    workbook = openpyxl.load_workbook('output.xlsx')
    # Select the worksheet
    worksheet = workbook.active
    # Select row 150 and fill columns A, B, C, D with some data
    worksheet.cell(row=row, column=3).value = positive_score
    worksheet.cell(row=row, column=4).value = -(negative_score)
    worksheet.cell(row=row, column=5).value = polarity_score
    worksheet.cell(row=row, column=6).value = subjectivity_score
    worksheet.cell(row=row, column=7).value = average_sentence_length
    worksheet.cell(row=row, column=8).value = percentage_complex_words
    worksheet.cell(row=row, column=9).value = fog_index
    worksheet.cell(row=row, column=10).value = avg_num_words_per_sentence
    worksheet.cell(row=row, column=11).value = complex_words
    worksheet.cell(row=row, column=12).value = total_words
    worksheet.cell(row=row, column=13).value = avg_syllables_per_word
    worksheet.cell(row=row, column=14).value = personal_pronouns
    worksheet.cell(row=row, column=15).value = avg_word_length

    # Save the changes
    workbook.save('output.xlsx')
    # Return dictionary of sentiment scores and readability score
    # I AM JUST RETURNING THIS VALUES FOR NO REASONS
    return {
        'positive_score': positive_score,
        'negative_score': -(negative_score),
        'polarity_score': polarity_score,
        'subjectivity_score': subjectivity_score,
        'avg_sentence_length':average_sentence_length,
        'percentage_complex_words':percentage_complex_words,
        'fog_index': fog_index,
        'avg_num_words_per_sentence':avg_num_words_per_sentence,
        'complex_words': complex_words,
        'total_words_count': total_words,
        'avg_syllables_per_word': avg_syllables_per_word,
        'personal_pronouns': personal_pronouns,
        'avg_word_length':avg_word_length,
    }

# open the folder that contains all extracted text from the article with their id as filename
folder_path = os.getcwd()+'\\ext_txt\\'
num = 0
row =2
for filename in os.listdir(folder_path):
    num +=1
    if filename.endswith(".txt"):
        file_path = os.path.join(folder_path, filename)
        with open(file_path, 'r', encoding='utf-8') as text_file:
            url_id = int(filename.split('.')[0])
            text = text_file.read()
            sentiment_scores = analyze_sentiment(text,url_id=url_id)

            # Print the sentiment scores
            # PRINTING VALUES TO CHECK, IN TERMINAL
            print(f'<<<<<<<<<<<<<<<<<<<<<<<<<<ARTICLE NO. {num} FileName {filename}>>>>>>>>>>>>>>>>>>>>>>>>>>>\n\n')
            print(sentiment_scores,'\n\n\n\n\n')